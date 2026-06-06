import csv
from io import BytesIO, StringIO
from datetime import datetime

from flask import Blueprint, request, send_file, jsonify
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .auth.auth_utils import login_required
from ..db.session import SessionLocal
from ..utils.report.report_context import make_region_slug

report_bp = Blueprint("report_bp", __name__)

# ── ID mappings — must match DB ────────────────────────────────────────────────
_HAZARD_ALIAS  = {"multi": "multihazard"}
_HAZARD_ID     = {"flood": 1, "drought": 2, "multihazard": 3}
_SCENARIO_ID   = {"nonclimate": 1, "climate": 2}
_CLIMATE_LABEL = {"nonclimate": "Baseline", "climate": "Projection"}
_RP_STR_TO_INT = {"rp25": 25, "rp50": 50, "rp100": 100, "rp250": 250}
_RP_ID         = {25: 1, 50: 2, 100: 3, 250: 4}
_HAZARD_LABEL  = {"flood": "Flood", "drought": "Drought", "multihazard": "Multi-hazard"}

ALLOWED_HAZARDS   = {"flood", "drought", "multi"}
ALLOWED_SCENARIOS = {"rp25", "rp50", "rp100", "rp250"}
ALLOWED_CLIMATE   = {"nonclimate", "climate"}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _validate(hazard, scenario, climate):
    if hazard not in ALLOWED_HAZARDS:
        return jsonify({"error": f"Hazard tidak valid: {hazard}"}), 400
    if scenario not in ALLOWED_SCENARIOS:
        return jsonify({"error": f"Scenario tidak valid: {scenario}"}), 400
    if climate not in ALLOWED_CLIMATE:
        return jsonify({"error": f"Climate tidak valid: {climate}"}), 400
    return None


def _get_ids(hazard, scenario, climate):
    db_hazard   = _HAZARD_ALIAS.get(hazard, hazard)
    hazard_id   = _HAZARD_ID[db_hazard]
    scenario_id = _SCENARIO_ID[climate]
    rp          = _RP_STR_TO_INT[scenario]
    rp_id       = _RP_ID[rp]
    return db_hazard, hazard_id, scenario_id, rp_id


def _latest_run_id(db):
    row = db.execute(text(
        "SELECT id FROM runs ORDER BY id DESC LIMIT 1"
    )).fetchone()
    return row.id if row else None


def _query_data(db, hazard_id, scenario_id, rp_id, run_id):
    """All regions with loss, aal, hazard_index, production for the given filters."""
    return db.execute(text("""
        SELECT
            r.id_kabkota,
            r.kab_kota,
            r.prov,
            COALESCE(l.loss,       0)::float AS loss,
            COALESCE(a.aal,        0)::float AS aal,
            COALESCE(z.mean_value, 0)::float AS hazard_index,
            COALESCE(p.total_prod, 0)::float AS total_prod
        FROM regions_adm r
        LEFT JOIN losses l
            ON  r.id_kabkota  = l.id_kabkota
            AND l.hazard_id   = :hazard_id
            AND l.scenario_id = :scenario_id
            AND l.rp_id       = :rp_id
            AND l.run_id      = :run_id
        LEFT JOIN aal a
            ON  r.id_kabkota  = a.id_kabkota
            AND a.hazard_id   = :hazard_id
            AND a.scenario_id = :scenario_id
            AND a.run_id      = :run_id
        LEFT JOIN zonal_kabupaten z
            ON  r.id_kabkota  = z.id_kabkota
            AND z.hazard_id   = :hazard_id
            AND z.scenario_id = :scenario_id
            AND z.rp_id       = :rp_id
            AND z.run_id      = :run_id
        LEFT JOIN (
            SELECT id_kabkota, SUM(total_prod)::float AS total_prod
            FROM production GROUP BY id_kabkota
        ) p ON r.id_kabkota = p.id_kabkota
        ORDER BY loss DESC NULLS LAST, r.kab_kota ASC
    """), {
        "hazard_id":   hazard_id,
        "scenario_id": scenario_id,
        "rp_id":       rp_id,
        "run_id":      run_id,
    }).fetchall()


def _aal_total(db, hazard_id, scenario_id, run_id, region: str = "", province: str = ""):
    if region:
        row = db.execute(text("""
            SELECT COALESCE(SUM(a.aal), 0)::float AS total
            FROM aal a
            JOIN regions_adm r ON a.id_kabkota = r.id_kabkota
            WHERE a.hazard_id   = :hid
              AND a.scenario_id = :sid
              AND a.run_id      = :run_id
              AND LOWER(TRIM(r.kab_kota)) = LOWER(TRIM(:region))
        """), {"hid": hazard_id, "sid": scenario_id, "run_id": run_id, "region": region}).fetchone()
    elif province:
        row = db.execute(text("""
            SELECT COALESCE(SUM(a.aal), 0)::float AS total
            FROM aal a
            JOIN regions_adm r ON a.id_kabkota = r.id_kabkota
            WHERE a.hazard_id   = :hid
              AND a.scenario_id = :sid
              AND a.run_id      = :run_id
              AND LOWER(TRIM(r.prov)) = LOWER(TRIM(:province))
        """), {"hid": hazard_id, "sid": scenario_id, "run_id": run_id, "province": province}).fetchone()
    else:
        row = db.execute(text("""
            SELECT COALESCE(SUM(aal), 0)::float AS total
            FROM aal
            WHERE hazard_id   = :hid
              AND scenario_id = :sid
              AND run_id      = :run_id
        """), {"hid": hazard_id, "sid": scenario_id, "run_id": run_id}).fetchone()
    return float(row.total) if row else 0.0


def _fmt(v):
    """Compact Rupiah formatter for summary cells."""
    try:
        v = float(v)
    except Exception:
        return "Rp 0"
    if abs(v) >= 1e12:
        return f"Rp {v / 1e12:.1f} T"
    if abs(v) >= 1e9:
        return f"Rp {v / 1e9:.1f} M"
    if abs(v) >= 1e6:
        return f"Rp {v / 1e6:.1f} Jt"
    return "Rp {:,.0f}".format(v).replace(",", ".")


# ══════════════════════════════════════════════════════════════════════════════
# CSV DOWNLOAD  (unchanged)
# ══════════════════════════════════════════════════════════════════════════════

@report_bp.route("/api/download-csv")
@login_required
def download_csv():
    hazard   = request.args.get("hazard",   "multi")
    scenario = request.args.get("scenario", "rp25")
    climate  = request.args.get("climate",  "nonclimate")
    region   = request.args.get("region",   "").strip()
    province = request.args.get("province", "").strip()
    run_id_param = request.args.get("run_id", type=int)

    err = _validate(hazard, scenario, climate)
    if err:
        return err

    _db_hazard, hazard_id, scenario_id, rp_id = _get_ids(hazard, scenario, climate)

    db = SessionLocal()
    try:
        run_id = run_id_param or _latest_run_id(db)
        if not run_id:
            return jsonify({"error": "No runs found"}), 404

        all_rows = _query_data(db, hazard_id, scenario_id, rp_id, run_id)

        # Compute national total BEFORE filtering so share % reflects national context
        total_loss = sum(r.loss or 0 for r in all_rows)

        if region:
            rows = [r for r in all_rows if r.kab_kota.strip().lower() == region.lower()]
        elif province:
            rows = [r for r in all_rows if r.prov.strip().lower() == province.lower()]
        else:
            rows = all_rows

        if not rows:
            return jsonify({"error": "Tidak ada data untuk filter yang dipilih."}), 404

        out = StringIO()
        writer = csv.writer(out)
        climate_label = _CLIMATE_LABEL.get(climate, climate)
        writer.writerow([
            "ID Kabkota",
            "Kabupaten / Kota",
            "Provinsi",
            f"Loss (Rp) — {hazard.upper()} {scenario.upper()} {climate_label}",
            f"AAL (Rp) — {hazard.upper()} {climate_label}",
            "Hazard Index (0–1)",
            "Total Produksi (ton)",
            "Persentase Kontribusi (%)",
        ])
        for r in rows:
            loss = r.loss or 0
            share = round((loss / total_loss) * 100, 2) if total_loss else 0.00
            writer.writerow([
                r.id_kabkota,
                r.kab_kota,
                r.prov,
                f"{loss:.2f}",
                f"{(r.aal or 0):.2f}",
                round(r.hazard_index, 6),
                round(r.total_prod, 2),
                f"{share:.2f}",
            ])

        buf = BytesIO(out.getvalue().encode("utf-8-sig"))  # UTF-8 BOM for Excel
        region_slug = make_region_slug(region or province)
        return send_file(
            buf,
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"padis_{hazard}_{climate_label.lower()}_{scenario}_{region_slug}.csv",
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# XLSX BUILDER — openpyxl
# ══════════════════════════════════════════════════════════════════════════════

# ── Style primitives ──────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex6.lstrip("#"))

def _font(bold=False, size=10, color="1F2937", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color.lstrip("#"), italic=italic)

def _border(style="thin", color="D1D5DB") -> Border:
    s = Side(style=style, color=color.lstrip("#"))
    return Border(left=s, right=s, top=s, bottom=s)

def _thin_bottom(color="D1D5DB") -> Border:
    s = Side(style="thin", color=color.lstrip("#"))
    return Border(bottom=s)

_FILL_NAVY    = _fill("0D2137")
_FILL_GOLD    = _fill("C9A227")
_FILL_BLUE    = _fill("1E63B5")
_FILL_LIGHT   = _fill("EFF6FF")
_FILL_ALT     = _fill("F9FAFB")
_FILL_WHITE   = _fill("FFFFFF")
_FILL_TOTAL   = _fill("F3F4F6")

_FONT_WHITE_B = _font(bold=True, color="FFFFFF")
_FONT_WHITE   = _font(color="FFFFFF")
_FONT_NAVY_B  = _font(bold=True, color="0D2137", size=9)
_FONT_DARK_B  = _font(bold=True, color="1F2937", size=9)
_FONT_DARK    = _font(color="1F2937", size=9)
_FONT_MUTED   = _font(color="6B7280", size=8, italic=True)
_FONT_LABEL   = _font(bold=True, color="6B7280", size=8)

_ALIGN_C  = Alignment(horizontal="center", vertical="center")
_ALIGN_R  = Alignment(horizontal="right",  vertical="center")
_ALIGN_L  = Alignment(horizontal="left",   vertical="center")
_ALIGN_LW = Alignment(horizontal="left",   vertical="top", wrap_text=True)

_BORDER_ALL  = _border()
_BORDER_NONE = Border()

_FMT_IDR  = '#,##0'     # raw integer, thousands-separated
_FMT_DEC  = '0.0000'    # 4-decimal for hazard index
_FMT_TON  = '#,##0'
_FMT_PCT  = '0.0"%"'    # e.g. 12.3%


def _set(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    """Write a cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:    cell.font      = font
    if fill:    cell.fill      = fill
    if align:   cell.alignment = align
    if border:  cell.border    = border
    if num_fmt: cell.number_format = num_fmt
    return cell


def _merge_row(ws, row, col_start, col_end, value, font=None, fill=None, align=None):
    """Merge cells across a row and write styled value."""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    # Apply fill to all merged cells (openpyxl only styles the top-left)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = fill or _FILL_WHITE
    return cell


def _col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ── Sheet 1: Ringkasan ────────────────────────────────────────────────────────

def _sheet_ringkasan(ws, ctx, all_rows):
    """Summary sheet: metadata, KPIs, top-3 highlights, insight narrative."""
    ws.sheet_view.showGridLines = False

    _col_widths(ws, {"A": 24, "B": 26, "C": 22, "D": 26, "E": 16, "F": 16})

    r = 1

    # ── Title banner ──────────────────────────────────────────────────────────
    _merge_row(ws, r, 1, 6,
               "PADIS — Laporan Analisis Risiko Bencana Pertanian",
               font=_font(bold=True, size=14, color="FFFFFF"),
               fill=_FILL_NAVY, align=_ALIGN_C)
    ws.row_dimensions[r].height = 28
    r += 1

    _merge_row(ws, r, 1, 6,
               "Paddy Disaster Information System · Teknik Geodesi dan Geomatika",
               font=_font(size=9, color="BFDBFE"),
               fill=_FILL_NAVY, align=_ALIGN_C)
    ws.row_dimensions[r].height = 16
    r += 1

    # Gold stripe row
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = _FILL_GOLD
    ws.row_dimensions[r].height = 3
    r += 1

    r += 1  # spacer

    # ── Metadata section ──────────────────────────────────────────────────────
    _merge_row(ws, r, 1, 6, "IDENTITAS LAPORAN",
               font=_font(bold=True, size=9, color="FFFFFF"),
               fill=_FILL_BLUE, align=_ALIGN_L)
    ws.row_dimensions[r].height = 18
    r += 1

    meta_pairs = [
        ("Jenis Bahaya",    ctx["hazard_label"],    "Skenario Iklim",  ctx["climate_label"]),
        ("Periode Ulang",   ctx["scenario_label"],  "Wilayah",         ctx["region"] or "Seluruh Indonesia"),
        ("Tanggal Dibuat",  ctx["generated_at"],    "Run ID",          f"#{ctx['run_id']}"),
    ]
    for label1, val1, label2, val2 in meta_pairs:
        _set(ws, r, 1, label1, font=_FONT_LABEL, fill=_FILL_LIGHT, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 2, val1,   font=_FONT_DARK_B, fill=_FILL_WHITE, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 3, label2, font=_FONT_LABEL, fill=_FILL_LIGHT, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 4, val2,   font=_FONT_DARK_B, fill=_FILL_WHITE, align=_ALIGN_L, border=_BORDER_ALL)
        ws.cell(row=r, column=5).fill = _FILL_WHITE
        ws.cell(row=r, column=6).fill = _FILL_WHITE
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1  # spacer

    # ── KPI section ───────────────────────────────────────────────────────────
    _merge_row(ws, r, 1, 6, "INDIKATOR UTAMA (KEY PERFORMANCE INDICATORS)",
               font=_font(bold=True, size=9, color="FFFFFF"),
               fill=_FILL_NAVY, align=_ALIGN_L)
    ws.row_dimensions[r].height = 18
    r += 1

    kpis = [
        ("Total Kerugian",       ctx["total_loss"],       "Kerugian Tertinggi",    ctx["top_loss"]),
        ("AAL Baseline",         ctx["aal_nonclimate"],   "AAL Projection",        ctx["aal_climate"]),
        ("Δ AAL (abs)",          ctx["aal_delta"],        "% Perubahan AAL",       ctx["aal_pct"]),
        ("Wilayah Terdampak",    ctx["valid_count"],      "Total Wilayah",         ctx["data_count"]),
        ("Top-3 Share (%)",      ctx["top3_share_label"], "Wilayah #1 Share (%)",  ctx["top1_share_label"]),
    ]
    for label1, val1, label2, val2 in kpis:
        _set(ws, r, 1, label1, font=_FONT_LABEL, fill=_FILL_LIGHT, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 2, val1,   font=_FONT_NAVY_B, fill=_FILL_WHITE, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 3, label2, font=_FONT_LABEL, fill=_FILL_LIGHT, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 4, val2,   font=_FONT_NAVY_B, fill=_FILL_WHITE, align=_ALIGN_L, border=_BORDER_ALL)
        ws.cell(row=r, column=5).fill = _FILL_WHITE
        ws.cell(row=r, column=6).fill = _FILL_WHITE
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1

    # ── Top 3 highlights ──────────────────────────────────────────────────────
    _merge_row(ws, r, 1, 6, "TIGA WILAYAH PRIORITAS RISIKO TERTINGGI",
               font=_font(bold=True, size=9, color="FFFFFF"),
               fill=_FILL_NAVY, align=_ALIGN_L)
    ws.row_dimensions[r].height = 18
    r += 1

    # Sub-header
    for col, (hdr, w) in enumerate(zip(
        ["No", "Kabupaten / Kota", "Provinsi", "Kerugian (Rp)", "Share (%)"],
        [4, 28, 20, 20, 10]
    ), start=1):
        _set(ws, r, col, hdr, font=_FONT_WHITE_B, fill=_FILL_BLUE,
             align=_ALIGN_C, border=_BORDER_ALL)
    ws.row_dimensions[r].height = 15
    r += 1

    top3 = ctx["top_regions"][:3]
    for i, row_data in enumerate(top3):
        fill = _FILL_ALT if i % 2 == 0 else _FILL_WHITE
        _set(ws, r, 1, i + 1,               font=_FONT_DARK,   fill=fill, align=_ALIGN_C, border=_BORDER_ALL)
        _set(ws, r, 2, row_data["kab_kota"],font=_FONT_DARK_B, fill=fill, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 3, row_data["prov"],     font=_FONT_DARK,   fill=fill, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 4, row_data["loss_fmt"], font=_FONT_DARK,   fill=fill, align=_ALIGN_R, border=_BORDER_ALL)
        _set(ws, r, 5, f'{row_data["share_pct"]}%', font=_FONT_DARK, fill=fill, align=_ALIGN_C, border=_BORDER_ALL)
        ws.row_dimensions[r].height = 14
        r += 1

    r += 1

    # ── Insight narrative ─────────────────────────────────────────────────────
    _merge_row(ws, r, 1, 6, "ANALISIS RINGKAS",
               font=_font(bold=True, size=9, color="FFFFFF"),
               fill=_FILL_NAVY, align=_ALIGN_L)
    ws.row_dimensions[r].height = 18
    r += 1

    insight = ctx.get("insight", "")
    _merge_row(ws, r, 1, 6, insight,
               font=_font(size=9, color="1F2937"),
               fill=_FILL_LIGHT, align=_ALIGN_LW)
    ws.row_dimensions[r].height = 48
    r += 1

    r += 1

    # ── Navigation note ───────────────────────────────────────────────────────
    note = ("File ini berisi 3 sheet:  "
            "[1] Ringkasan — metadata & KPI  |  "
            "[2] 10 Wilayah Teratas — top 10 kabupaten/kota  |  "
            "[3] Semua Data — seluruh wilayah teranalisis")
    _merge_row(ws, r, 1, 6, note,
               font=_font(size=8, color="6B7280", italic=True),
               fill=_FILL_WHITE, align=_ALIGN_LW)
    ws.row_dimensions[r].height = 28


# ── Sheet 2 & 3: Data tables ──────────────────────────────────────────────────

_DATA_HEADERS = [
    ("No",              5),
    ("ID Kabkota",      13),
    ("Kabupaten / Kota",32),
    ("Provinsi",        22),
    ("Loss (Rp)",       20),
    ("AAL (Rp)",        20),
    ("Hazard Index",    14),
    ("Produksi (ton)",  18),
    ("Share (%)",       11),
]


def _fill_data_sheet(ws, rows, total_loss, sheet_title):
    """Render a styled data table onto ws."""
    ws.sheet_view.showGridLines = False

    # Column widths
    for col, (_, w) in enumerate(_DATA_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    r = 1

    # Banner
    _merge_row(ws, r, 1, len(_DATA_HEADERS), sheet_title,
               font=_font(bold=True, size=12, color="FFFFFF"),
               fill=_FILL_NAVY, align=_ALIGN_C)
    ws.row_dimensions[r].height = 22
    r += 1

    # Gold stripe
    for c in range(1, len(_DATA_HEADERS) + 1):
        ws.cell(row=r, column=c).fill = _FILL_GOLD
    ws.row_dimensions[r].height = 3
    r += 1

    # Column headers
    for col, (hdr, _) in enumerate(_DATA_HEADERS, start=1):
        _set(ws, r, col, hdr, font=_FONT_WHITE_B, fill=_FILL_BLUE,
             align=_ALIGN_C, border=_BORDER_ALL)
    ws.row_dimensions[r].height = 16
    r += 1

    # Data rows
    for i, row_data in enumerate(rows):
        fill = _FILL_ALT if i % 2 == 0 else _FILL_WHITE
        share = (row_data.loss / total_loss * 100) if total_loss and row_data.loss else 0.0

        _set(ws, r, 1, i + 1,               font=_FONT_MUTED,  fill=fill, align=_ALIGN_C, border=_BORDER_ALL)
        _set(ws, r, 2, row_data.id_kabkota,  font=_FONT_DARK,   fill=fill, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 3, row_data.kab_kota,    font=_FONT_DARK_B, fill=fill, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 4, row_data.prov,        font=_FONT_DARK,   fill=fill, align=_ALIGN_L, border=_BORDER_ALL)
        _set(ws, r, 5, row_data.loss,        font=_FONT_DARK,   fill=fill, align=_ALIGN_R, border=_BORDER_ALL, num_fmt=_FMT_IDR)
        _set(ws, r, 6, row_data.aal,         font=_FONT_DARK,   fill=fill, align=_ALIGN_R, border=_BORDER_ALL, num_fmt=_FMT_IDR)
        _set(ws, r, 7, row_data.hazard_index,font=_FONT_DARK,   fill=fill, align=_ALIGN_C, border=_BORDER_ALL, num_fmt=_FMT_DEC)
        _set(ws, r, 8, row_data.total_prod,  font=_FONT_DARK,   fill=fill, align=_ALIGN_R, border=_BORDER_ALL, num_fmt=_FMT_TON)
        _set(ws, r, 9, round(share, 2),      font=_FONT_DARK,   fill=fill, align=_ALIGN_C, border=_BORDER_ALL, num_fmt=_FMT_PCT)
        ws.row_dimensions[r].height = 13
        r += 1

    # Total footer row
    total_share = 100.0 if total_loss else 0.0
    _set(ws, r, 1, "",         fill=_FILL_TOTAL, border=_BORDER_ALL)
    _set(ws, r, 2, "",         fill=_FILL_TOTAL, border=_BORDER_ALL)
    _set(ws, r, 3, "TOTAL",    font=_font(bold=True, size=9, color="0D2137"), fill=_FILL_TOTAL, align=_ALIGN_L, border=_BORDER_ALL)
    _set(ws, r, 4, "",         fill=_FILL_TOTAL, border=_BORDER_ALL)
    _set(ws, r, 5, total_loss, font=_font(bold=True, size=9, color="0D2137"), fill=_FILL_TOTAL, align=_ALIGN_R, border=_BORDER_ALL, num_fmt=_FMT_IDR)
    _set(ws, r, 6, "",         fill=_FILL_TOTAL, border=_BORDER_ALL)
    _set(ws, r, 7, "",         fill=_FILL_TOTAL, border=_BORDER_ALL)
    _set(ws, r, 8, "",         fill=_FILL_TOTAL, border=_BORDER_ALL)
    _set(ws, r, 9, round(total_share, 1), font=_font(bold=True, size=9, color="0D2137"), fill=_FILL_TOTAL, align=_ALIGN_C, border=_BORDER_ALL, num_fmt=_FMT_PCT)
    ws.row_dimensions[r].height = 15

    # Freeze header rows
    ws.freeze_panes = ws.cell(row=4, column=1)


def _build_xlsx(buf, ctx, all_rows):
    """Render multi-sheet XLSX into buf."""
    wb = Workbook()

    # Sheet 1 — Ringkasan
    ws_sum = wb.active
    ws_sum.title = "Ringkasan"
    _sheet_ringkasan(ws_sum, ctx, all_rows)

    # Sheet 2 — Top 10 Wilayah
    ws_top = wb.create_sheet("10 Wilayah Teratas")
    top10 = [r for r in all_rows if r.loss > 0][:10]
    total_loss = sum(r.loss for r in all_rows if r.loss)
    _fill_data_sheet(
        ws_top, top10, total_loss,
        f"10 Wilayah Kerugian Tertinggi — {ctx['hazard_label']} · {ctx['scenario_label']} · {ctx['climate_label']}"
    )

    # Sheet 3 — Semua Data
    ws_all = wb.create_sheet("Semua Data")
    _fill_data_sheet(
        ws_all, all_rows, total_loss,
        f"Semua Data Wilayah — {ctx['hazard_label']} · {ctx['scenario_label']} · {ctx['climate_label']}"
    )

    wb.save(buf)


# ══════════════════════════════════════════════════════════════════════════════
# XLSX REPORT ENDPOINT
# ══════════════════════════════════════════════════════════════════════════════

@report_bp.route("/api/generate-report-v2")
@login_required
def generate_report_v2():
    hazard   = request.args.get("hazard",   "multi")
    scenario = request.args.get("scenario", "rp25")
    climate  = request.args.get("climate",  "nonclimate")
    region   = request.args.get("region",   "").strip()
    province = request.args.get("province", "").strip()
    run_id_param = request.args.get("run_id", type=int)

    err = _validate(hazard, scenario, climate)
    if err:
        return err

    db_hazard, hazard_id, scenario_id, rp_id = _get_ids(hazard, scenario, climate)

    db = SessionLocal()
    try:
        run_id = run_id_param or _latest_run_id(db)
        if not run_id:
            return jsonify({"error": "No runs found"}), 404

        rows = _query_data(db, hazard_id, scenario_id, rp_id, run_id)

        if region:
            rows = [r for r in rows if r.kab_kota.strip().lower() == region.lower()]
        elif province:
            rows = [r for r in rows if r.prov.strip().lower() == province.lower()]

        # ── Statistics ────────────────────────────────────────────────────────
        valid       = [r for r in rows if r.loss > 0]
        data_count  = len(rows)
        valid_count = len(valid)
        total_loss  = sum(r.loss for r in valid)
        top_rows    = valid[:10]

        aal_nc    = _aal_total(db, hazard_id, _SCENARIO_ID["nonclimate"], run_id, region, province)
        aal_cc    = _aal_total(db, hazard_id, _SCENARIO_ID["climate"],    run_id, region, province)
        aal_delta = aal_cc - aal_nc
        aal_pct   = ((aal_delta / aal_nc) * 100.0) if aal_nc else 0.0

        top_row        = valid[0] if valid else None
        top_name       = f"{top_row.kab_kota}, {top_row.prov}" if top_row else "-"
        top_loss_v     = top_row.loss if top_row else 0.0
        top_loss_share = (top_loss_v / total_loss * 100.0) if total_loss else 0.0
        top3_loss      = sum(r.loss for r in valid[:3])
        top3_share     = (top3_loss / total_loss * 100.0) if total_loss else 0.0

        formatted_top = [
            {
                "rank":      i + 1,
                "kab_kota":  r.kab_kota,
                "prov":      r.prov,
                "loss_fmt":  _fmt(r.loss),
                "aal_fmt":   _fmt(r.aal),
                "hidx":      f"{r.hazard_index:.4f}",
                "prod_fmt":  f"{r.total_prod:,.0f}".replace(",", "."),
                "share_pct": round((r.loss / total_loss * 100) if total_loss else 0, 1),
            }
            for i, r in enumerate(top_rows)
        ]

        # Narrative insight for Summary sheet
        insight_parts = [
            f"Berdasarkan analisis {_HAZARD_LABEL.get(db_hazard, db_hazard)} skenario "
            f"{'Projection' if climate == 'climate' else 'Baseline'} "
            f"periode ulang {scenario.upper()}, total kerugian produksi padi mencapai "
            f"{_fmt(total_loss)} dari {valid_count} kabupaten/kota terdampak "
            f"(dari {data_count} wilayah teranalisis)."
        ]
        if top_row:
            insight_parts.append(
                f"Wilayah tertinggi adalah {top_name} ({top_loss_share:.1f}% dari total)."
            )
        if len(valid) >= 3:
            insight_parts.append(
                f"Tiga wilayah teratas menyumbang {top3_share:.1f}% dari total kerugian."
            )
        if aal_nc:
            arrow = "meningkat" if aal_delta >= 0 else "menurun"
            insight_parts.append(
                f"Proyeksi AAL {arrow} {abs(aal_pct):.1f}% pada skenario perubahan iklim "
                f"({_fmt(aal_nc)} → {_fmt(aal_cc)})."
            )

        ctx = {
            "hazard_label":      _HAZARD_LABEL.get(db_hazard, db_hazard),
            "climate_label":     "Projection" if climate == "climate" else "Baseline",
            "scenario_label":    scenario.upper(),
            "generated_at":      datetime.now().strftime("%d %B %Y, %H:%M WIB"),
            "region":            region or province,
            "is_regional":       bool(region or province),
            "run_id":            run_id,

            "data_count":        data_count,
            "valid_count":       valid_count,
            "total_loss":        _fmt(total_loss),
            "top_region":        top_name,
            "top_loss":          _fmt(top_loss_v),
            "top1_share_label":  f"{top_loss_share:.1f}%",
            "top3_share_label":  f"{top3_share:.1f}%",

            "aal_nonclimate":    _fmt(aal_nc),
            "aal_climate":       _fmt(aal_cc),
            "aal_delta":         _fmt(abs(aal_delta)),
            "aal_pct":           f"{aal_pct:+.1f}%",
            "aal_pct_up":        aal_delta >= 0,

            "top_regions":       formatted_top,
            "empty_state":       valid_count == 0,
            "insight":           " ".join(insight_parts),
        }
    finally:
        db.close()

    buf = BytesIO()
    _build_xlsx(buf, ctx, rows)
    buf.seek(0)

    region_slug = make_region_slug(region or province)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"padis_data_{hazard}_{climate}_{scenario}_{region_slug}.xlsx",
    )
